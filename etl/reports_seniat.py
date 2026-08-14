"""
reports_seniat.py — Generates SENIAT-format fiscal books as XLSX: Libro de
Ventas, Libro de Compras, and IVA/ISLR retention summaries.

Reads directly from the Profit Plus source DB (not DW_Profit) — these are
point-in-time regulatory exports for a given month, not warehouse data.

Usage:
    python etl/reports_seniat.py --period 202606 --out reports/
    # writes reports/libro_ventas_202606.xlsx, libro_compras_202606.xlsx,
    #        retenciones_iva_202606.xlsx, retenciones_islr_202606.xlsx

Business rules (see erp-knowledge-base/docs/WORKFLOWS.md):
  - anulado = 0 always filtered for reporting.
  - USD conversion uses the DOCUMENT's own tasa, never the current rate.
  - IVA retention detail lives in saCobroRetenIvaReng (retenido a nosotros al
    cobrar) / saPagoRetenIvaReng (retenido por nosotros al pagar), linked via
    rowguid_reng_cob -> saCobroDocReng.rowguid / saPagoDocReng.rowguid.
  - ISLR retention detail lives in saCobroRentenReng / saPagoRentenReng,
    same join pattern via rowguid_reng_cob.
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from etl._db import load_dotenv, get_connections  # noqa: E402

try:
    from openpyxl import Workbook
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")


LIBRO_VENTAS_SQL = """
SELECT
    f.fec_emis, f.n_control, f.doc_num, c.co_cli, c.cli_des, c.rif,
    f.total_bruto, f.monto_imp, f.total_neto, f.tasa,
    f.total_neto / NULLIF(f.tasa, 0) AS total_neto_usd
FROM saFacturaVenta f
LEFT JOIN saCliente c ON f.co_cli = c.co_cli
WHERE f.anulado = 0
  AND f.fec_emis >= ? AND f.fec_emis < ?
ORDER BY f.fec_emis, f.doc_num;
"""

LIBRO_COMPRAS_SQL = """
SELECT
    f.fec_emis, f.n_control, f.doc_num, p.co_prov, p.prov_des, p.rif,
    f.total_bruto, f.monto_imp, f.total_neto, f.tasa,
    f.total_neto / NULLIF(f.tasa, 0) AS total_neto_usd
FROM saFacturaCompra f
LEFT JOIN saProveedor p ON f.co_prov = p.co_prov
WHERE f.anulado = 0
  AND f.fec_emis >= ? AND f.fec_emis < ?
ORDER BY f.fec_emis, f.doc_num;
"""

RETENCIONES_IVA_COBROS_SQL = """
SELECT
    r.periodo_impositivo, r.rif_contribuyente, r.numero_documento, r.numero_control_documento,
    r.fecha_documento, r.tipo_documento, r.monto_documento, r.base_imponible,
    r.alicuota, r.monto_ret_imp, r.num_comprobante
FROM saCobroRetenIvaReng r
INNER JOIN saCobroDocReng cr ON r.rowguid_reng_cob = cr.rowguid
INNER JOIN saCobro c ON cr.cob_num = c.cob_num
WHERE c.anulado = 0 AND r.periodo_impositivo = ?
ORDER BY r.fecha_documento, r.numero_documento;
"""

RETENCIONES_IVA_PAGOS_SQL = """
SELECT
    r.periodo_impositivo, r.rif_contribuyente, r.numero_documento, r.numero_control_documento,
    r.fecha_documento, r.tipo_documento, r.monto_documento, r.base_imponible,
    r.alicuota, r.monto_ret_imp, r.num_comprobante
FROM saPagoRetenIvaReng r
INNER JOIN saPagoDocReng pr ON r.rowguid_reng_cob = pr.rowguid
INNER JOIN saPago p ON pr.cob_num = p.cob_num
WHERE p.anulado = 0 AND r.periodo_impositivo = ?
ORDER BY r.fecha_documento, r.numero_documento;
"""

RETENCIONES_ISLR_PAGOS_SQL = """
SELECT
    p.fecha, pr.nro_doc, pr.nro_fact, r.co_islr, r.monto, r.monto_reten,
    r.monto_obj, r.porc_retn
FROM saPagoRentenReng r
INNER JOIN saPagoDocReng pr ON r.rowguid_reng_cob = pr.rowguid
INNER JOIN saPago p ON pr.cob_num = p.cob_num
WHERE p.anulado = 0
  AND p.fecha >= ? AND p.fecha < ?
ORDER BY p.fecha, pr.nro_doc;
"""

RETENCIONES_ISLR_COBROS_SQL = """
SELECT
    c.fecha, cr.nro_doc, r.co_islr, r.monto, r.monto_reten,
    r.monto_obj, r.porc_retn
FROM saCobroRentenReng r
INNER JOIN saCobroDocReng cr ON r.rowguid_reng_cob = cr.rowguid
INNER JOIN saCobro c ON cr.cob_num = c.cob_num
WHERE c.anulado = 0
  AND c.fecha >= ? AND c.fecha < ?
ORDER BY c.fecha, cr.nro_doc;
"""


def _period_bounds(period: str):
    """period is YYYYMM -> (first day of month, first day of next month) as date strings."""
    year, month = int(period[:4]), int(period[4:6])
    start = f"{year:04d}-{month:02d}-01"
    if month == 12:
        end = f"{year + 1:04d}-01-01"
    else:
        end = f"{year:04d}-{month + 1:02d}-01"
    return start, end


def _write_xlsx(headers, rows, out_path: Path) -> int:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(rows)


def _libro_ventas(cur, start: str, end: str, out_dir: Path, period: str) -> int:
    cur.execute(LIBRO_VENTAS_SQL, (start, end))
    rows = cur.fetchall()
    headers = ["Fecha Emision", "N Control", "Num Factura", "Cod Cliente", "Nombre Cliente",
               "RIF", "Total Bruto", "Monto IVA", "Total Neto BS", "Tasa", "Total Neto USD"]
    return _write_xlsx(headers, rows, out_dir / f"libro_ventas_{period}.xlsx")


def _libro_compras(cur, start: str, end: str, out_dir: Path, period: str) -> int:
    cur.execute(LIBRO_COMPRAS_SQL, (start, end))
    rows = cur.fetchall()
    headers = ["Fecha Emision", "N Control", "Num Factura", "Cod Proveedor", "Nombre Proveedor",
               "RIF", "Total Bruto", "Monto IVA", "Total Neto BS", "Tasa", "Total Neto USD"]
    return _write_xlsx(headers, rows, out_dir / f"libro_compras_{period}.xlsx")


def _retenciones_iva(cur, period: str, out_dir: Path) -> int:
    headers = ["Periodo", "RIF Agente Retencion", "Num Documento", "Num Control",
               "Fecha Documento", "Tipo Documento", "Monto Documento", "Base Imponible",
               "Alicuota", "IVA Retenido", "Num Comprobante"]

    cur.execute(RETENCIONES_IVA_COBROS_SQL, (int(period),))
    cobros = cur.fetchall()
    cur.execute(RETENCIONES_IVA_PAGOS_SQL, (int(period),))
    pagos = cur.fetchall()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "IVA Retenido por Clientes"
    ws1.append(headers)
    for row in cobros:
        ws1.append(list(row))

    ws2 = wb.create_sheet("IVA Retenido a Proveedores")
    ws2.append(headers)
    for row in pagos:
        ws2.append(list(row))

    out_path = out_dir / f"retenciones_iva_{period}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(cobros) + len(pagos)


def _retenciones_islr(cur, start: str, end: str, period: str, out_dir: Path) -> int:
    headers_pagos = ["Fecha", "Num Documento", "Num Factura", "Cod ISLR", "Monto", "Monto Retenido",
                      "Monto Objeto", "Porcentaje"]
    headers_cobros = ["Fecha", "Num Documento", "Cod ISLR", "Monto", "Monto Retenido",
                       "Monto Objeto", "Porcentaje"]

    cur.execute(RETENCIONES_ISLR_PAGOS_SQL, (start, end))
    pagos = cur.fetchall()
    cur.execute(RETENCIONES_ISLR_COBROS_SQL, (start, end))
    cobros = cur.fetchall()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ISLR Retenido a Proveedores"
    ws1.append(headers_pagos)
    for row in pagos:
        ws1.append(list(row))

    ws2 = wb.create_sheet("ISLR Retenido por Clientes")
    ws2.append(headers_cobros)
    for row in cobros:
        ws2.append(list(row))

    out_path = out_dir / f"retenciones_islr_{period}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(pagos) + len(cobros)


def main() -> None:
    load_dotenv()

    parser = argparse.ArgumentParser(description="Generate SENIAT fiscal books (XLSX) for a given period")
    parser.add_argument("--period", required=True, help="Period YYYYMM, e.g. 202606")
    parser.add_argument("--out", default="reports", help="Output directory (default: reports/)")
    args = parser.parse_args()

    start, end = _period_bounds(args.period)
    out_dir = Path(args.out)

    print(f"[*] SENIAT reports for period {args.period} ({start} to {end})")
    src_conn, _ = get_connections(ensure_dw=False)
    cur = src_conn.cursor()

    try:
        n = _libro_ventas(cur, start, end, out_dir, args.period)
        print(f"  [+] libro_ventas_{args.period}.xlsx: {n} rows")

        n = _libro_compras(cur, start, end, out_dir, args.period)
        print(f"  [+] libro_compras_{args.period}.xlsx: {n} rows")

        n = _retenciones_iva(cur, args.period, out_dir)
        print(f"  [+] retenciones_iva_{args.period}.xlsx: {n} rows")

        n = _retenciones_islr(cur, start, end, args.period, out_dir)
        print(f"  [+] retenciones_islr_{args.period}.xlsx: {n} rows")

        print(f"\n[✔] SENIAT reports written to {out_dir}/")
    finally:
        src_conn.close()


if __name__ == "__main__":
    main()
